import sys
import os
import json
import openpyxl as xl
import logging



class File_Manager():
    """ Class to handle read and write operations on files """
    def __init__(self,directorypath):
        self.directory=directorypath

    def get_file_list(self):
        try:
            files = [f for f in os.listdir(self.directory) if os.path.isfile(os.path.join(self.directory,f))]
        except FileNotFoundError as ex:
            logging.critical("Directory not found : %s",ex.__str__ )
        except PermissionError as exp:     
            logging.critical("Directory Permission denied : %s",exp.__str__ )
        return files
    
    def delete_duplicate_file(self,file_list):
        """ Delete duplicate file """
        data_set = set(file_list)
        try:
            if data_set in file_list:
                print(f"trying to delete: {data_set}")
                os.remove(self.directory+data_set)
        except FileNotFoundError as ferror:
            return ferror
        except PermissionError as exp:
            logging.critical("Directory Permission denied : %s ",exp.__str__ )
            return exp
            
        
    def read_file(self,filename):
        """ Read file content"""
        try:
            with open(filename,'r',encoding="UTF-8") as file:
                data1 = file.read()
                return data1
        except FileNotFoundError as ferror:   
            return ferror 
    def read_file_per_line(self,filename):
        """ Read file content line by line """
        try:
            with open(filename,'r',encoding="UTF-8") as file:
                data1 = file.readlines()
                return data1
        except FileNotFoundError as ferror:   
            return ferror      
    
    def write_file(self,filename,data):
        """ Write data to file """
        try:
            with open(filename,'w',encoding="UTF-8") as file:
                file.write(data)
        except FileNotFoundError as ferror:
            return ferror
        
    def append_to_file(self,filename,data):
        """Append data to existing file"""
        try:
            with open(filename,'a',encoding="UTF-8") as file:
                file.write(data)
        except FileNotFoundError as ferror:   
            return ferror
    def read_update_file(self,filename,data=""):
        try:
            with open(filename,'r+',encoding="UTF-8") as file:
                file.write(data)
        except FileNotFoundError as ferror:
            return ferror
    def write_update_file(self,filename,data=""):
        try:
            with open(filename,'w+',encoding="UTF-8") as file:
                file.write(data)
        except FileNotFoundError as ferror:
            return ferror
    
    
    def read_excel_file(self):
        """ Read excel file and return as a Matix i,j to explore data """
        try:
            book = xl.load_workbook(self.directory)
            sheet = book.active
            data = []
            for row in  range(1,sheet.max_row+1):
                row_data = []
                for col in range(1,sheet.max_column+1):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(cell.value)
                data.append(row_data)
            return data
        except FileNotFoundError as ferror:
            logging.critical("File not found : %s",ferror.__str__ )
        except PermissionError as exp:
            logging.critical("Directory Permission denied : %s",exp.__str__ )
            
    def read_json_file(self):
        """ Wrtie data into excel file"""
        try:
            with open(self.directory,"r",encoding="utf-8") as f:
                data_1 = json.loads(f)
                data_list = data_1["students"]
            return data_list
        except FileNotFoundError as file:
            return file
        

if __name__ == "__main__":
    
    file_json= File_Manager("C:\\Data\\formdata.json")
    data2 = file_json.read_json_file
    print(type(data2))
   
